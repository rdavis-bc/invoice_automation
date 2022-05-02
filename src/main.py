from __future__ import annotations
import abc
from ast import arg
from email.policy import default
from functools import total_ordering, wraps
from optparse import Option
from typing import Any, Callable, List, Mapping, Optional, Protocol, Union, Tuple, NamedTuple, cast
import weakref
import openpyxl
import datetime as dt
import os 
from pathlib import Path
import re
import xlwings as xw
import glob
from operator import itemgetter 
from dataclasses import dataclass
import sys
from excel_analytics import utils


def create_data_directory_check(function: Callable[..., Any]) -> Callable[..., Any]:
    # TODO: Check if robust to the passing of multiple positional arguments
    @wraps(function)
    def wrapped_function(*args, **kwargs):
        args_list = list(args)
        print(args_list)
        directory = f"{args_list[0] + '/' if len(args_list) > 0 and args_list[0] else ''}{dt.datetime.now().strftime('%Y-%m-%d')}"
        print(directory)
        path = Path(os.getcwd())
        print(path)
        dir_path = os.path.join(path, directory)
        print(dir_path)
        
        if not Path(dir_path).is_dir():
            #TODO: figure out why return has to be added here
            # It seems that the return value has to be brought back up through the stacks
            return function(*args, **kwargs)
        else:
            print(f'{dir_path} already exists skipping creation')
            return dir_path
    
    return wrapped_function

@create_data_directory_check
def create_data_directory(base_dir: Optional[str] = None) -> str:
    """
    A utility function to quickly create a new directory with the option
    to be placed in a base directory
    Params:
        base_dir: str, a relative path in which you want to place your folder
    Returns: the relative path to the created directory
    Raises:
    >>> results = create_data_directory(base_dir='base_dir')
    >>> results

    """
    try:
        the_day = dt.datetime.now().strftime('%Y-%m-%d')
        if not base_dir: #creating the directory name
            directory = f"{the_day}"
        else:
            directory = f"{base_dir}/{the_day}"
        path = Path(os.getcwd())
        dir_path = os.path.join(path, directory)
        
        os.makedirs(dir_path)
    except Exception as e:
        print(f'{directory} or {base_dir} already exists Error:{e!r}')
    print(dir_path)
    return dir_path

class UserInputFormatter:
    """
    This is to be used as decorator that will format inputted values from the user that are passed
    to a function as kwargs
    """
    def __init__(self, str_format_method: Optional[str]=None) -> None:
        self.str_format_method = str_format_method

    def __call__(self, function: Callable[..., Any]) -> Any:
        @wraps(function)
        def wrapped_function(*args, **kwargs)-> Any:
            
            if self.str_format_method in ('split', 'lower', 'title', 'upper'):
                for key in list(kwargs.keys()):

                    str_start_value = kwargs[key]
                    run_format = "\'" + str_start_value +  "\'"  + f'.{self.str_format_method}()'
                    str_formatted_value = eval(run_format)
                    # TODO: use the update function instead
                    kwargs[key] = str_formatted_value
                    print(f'Transforming str from {str_start_value} -> {str_start_value} at the following key: {key}')
                # TODO: allow the args to be updated while skipping the self argument for the method
                # args = []
                # for arg in args:
                #     if arg == eval(self):
                #         continue
                #     run_format = "\'" + arg +  "\'"  + f'.{self.str_format_method}()'
                #     str_formatted_value = eval(run_format)
                #     arg=str_formatted_value
                #     args.append(arg)
                
            return function(*args, **kwargs)
        
        return wrapped_function



class UserInput():
    all_inputs: list = []
    """
    Represents the user input
    >>> UserInput(base_dir=sys.argv[1], dest_dir=sys.argv[2])
    """
    @UserInputFormatter('lower')
    def __init__(self, base_dir: Optional[str], dest_dir: Optional[str]) -> None:
        self.base_dir = base_dir
        self.dest_dir = dest_dir
        self.all_inputs.append((self.base_dir, self.dest_dir,))

    def copying_file_to_dest(self) -> None :
        print(f'copying data to {self.dest_dir}')

class AdditionalUserInput(UserInput):
    """Inherits from UserInput which was used previously as a helper"""

    all_inputs: list = []
    def __init__(self, base_dir_and_dest_dir: Tuple[str, str]) -> None:
        super().__init__(*base_dir_and_dest_dir)

    def copying_file_to_dest(self) -> None: 
        super().copying_file_to_dest()
        print(f'copying data to {self.dest_dir} in {self.__class__}') 

    def logging_choices(self) -> None:
        self.all_inputs.append((self.base_dir, self.dest_dir,))
        print(f'{self.base_dir!r} and {self.dest_dir!r} logged')

class Observer(Protocol):
    def __init__(self) -> None:
        self._created_at : dt.datetime

    def __call__(self) -> Optional[Any]:
        ...

class SheetSerializerObserver(Observer):
    def __init__(self) -> None:
        self._user = Path(os.getcwd()).owner()
        self._name = 'Sheet_Serializer_Observer'
        self._created_at = dt.datetime.now()

    def __call__(self) -> Optional[Any]:
        ol = ObserverLogging()
        ol(self)
        with ol as ol_file:
            pass


class SheetSerializer(abc.ABC):
    
    def __init__(self, sheet,/, client_name:str, month:str, dest_dir: str, ext: str) -> None:
        self._ext : str
        self.sheet = sheet
        self.client_name = client_name
        self.month = month
        self.dest_dir = dest_dir
        self.observers: List[Observer] = []
        if not self.ext == ext:
            raise ValueError(f"Wrong file format, not a {self.ext}")
        
    @abc.abstractmethod
    def serialize(self) -> None:
        ... 

    @property
    @abc.abstractmethod
    def ext(self) -> str :
        ...

class TestSerializer(SheetSerializer):
    ext: str = '.parquet'

class MissingInputError(ValueError):
    def __init__(self, base_dir: Optional[str] = '_', dest_dir: Optional[str] = '_') -> None:
        super().__init__(f'cant convert to PDFs while we\'re missing one of these: base_dir-{base_dir}, dest_dir-{dest_dir}')


def SerializationLogger(serializer: Callable[[None],None]):
    @wraps(serializer)
    def wrapped_function(*args):
        write_file = 'logging_file.txt'
        write_file_path = os.path.join(os.getcwd(), write_file)
        mode = 'a' if Path(write_file_path).is_file() else 'w'
            
        with open('logging_file.txt', mode) as file:
            file.write('\n') if mode == 'a' else None
            file.write(f'serialization taking place at {dt.datetime.now()}')
        cast(PDFSerializer, args[0]).call_observers()
        serializer(*args)
        
    return wrapped_function

class PDFSerializer(SheetSerializer):

    def __init__(self, sheet, client_name: str, month: str, dest_dir: str, ext: str = '.pdf') -> None:
        self._ext = '.pdf'
        super().__init__(sheet, client_name, month, dest_dir, ext=ext)
    
    @SerializationLogger
    def serialize(self) -> None:
        global wkbk
        global directory
        xw_wkbk = xw.Book(f"{wkbk}")
        xw_sheet = xw_wkbk.sheets[self.sheet]
        pdf_path = f'{directory}/{self.client_name + self.month}{self.ext}'
        print(pdf_path)
        try:
            xw_wkbk.to_pdf(path=pdf_path, include=self.sheet)
        except Exception as e:
            print(e)
    
    def attach_observer(self, observer:Observer) -> None:
        self.observers.append(observer)

    def remove_observer(self, observer:Observer) -> None:
        self.observers.remove(observer)

    def call_observers(self) -> None:
        [observer() for observer in self.observers]

    @property
    def ext(self) -> str:
        return self._ext

class OnlyUpdateDictionary(dict):
    """
    Dictionary to be used with the default_arg_dict in WorkbookParser to only allow the values to be updated rather than
    New ones to be introduced 
    TODO: See if Kwargs can be allowed ,check whether input is list or a mapping
    """
    def update(self, new_mapping: Mapping[str, str]):
        if set(new_mapping.keys()).issubset(set(self.keys())) != True:
            raise ValueError
        super().update(new_mapping)
        return self

class WorkbookParser():
    default_arg_dict: dict[str, Any] = {
        'lazy_processing':False ,
        'internet_access':True
    }
    def __init__(self, workbook: openpyxl.Workbook, *sheetinstances: Optional["Worksheet"], **kwargs: str) -> None:
        self.wkbk = workbook
        self.wkshts : Optional[list[weakref.ReferenceType["Worksheet"]]] = [sheet for sheet in sheetinstances] if sheetinstances else None
        self.wkshts2 = [*sheetinstances] if sheetinstances else None
        self._current_iteration: int = 0
        self.arg_dict: dict[str, Any] = {**self.default_arg_dict, **kwargs}
        # self.clientname: Optional[str] = None
        # self.month: Optional[str] = None
    
    @property
    def current_iteration(self) -> int:
        """This property keeps track of the sheets that our self.parser method goes through"""
        return self._current_iteration
        
    @current_iteration.setter
    def current_iteration(self, ix: int) -> None:
        self._current_iteration += ix

    def parser(self) -> None:
        global anas_input
        sheets = self.wkbk.sheetnames[1:]
        for ix, sheet in enumerate(sheets):
            client = self.wkbk.worksheets[ix]
            mx_rw, mx_col = client.max_row, client.max_column
            prev_val: Optional[str] = None
            for row in range(1, mx_rw + 1): # starting with the first row (top)
                for col in range(mx_col, 0, -1): # starting with the last column
                    val = client.cell(row=row, column=col).value #value of the coordinates
                    lw_cs_val = val.lower() if type(val) == str else val # trnasofmring to lower case if datatype is str
                    if re.match(r'.*date.*', lw_cs_val if type(val) == str else 'negative') and type(prev_val) == dt.datetime :# if date is in the current value and the 
                        # cell_contents = prev_val.split(' ')                               # is of type datetime      # re.match(r'.*\d{4}
                        prev_val_month :dt.datetime = prev_val.strftime("%b")
                        # insert code for creating pdf
                        pdf_serializer = PDFSerializer(sheet=sheet, client_name=sheet, month=prev_val_month, dest_dir=anas_input.dest_dir)
                        sso = SheetSerializerObserver()
                        pdf_serializer.attach_observer(sso)
                        pdf_serializer.serialize()
                        prev_val = None
                        self.current_iteration = 1
                        break
                    prev_val = val
                    self.current_iteration = 1

class Input(NamedTuple):
    staging_dir:str
    destination_dir: str




@total_ordering
@dataclass(frozen=True)
class AlternativeInput:
    staging_dir:str
    destination_dir: str
    current_time: dt.datetime = dt.datetime.now()
    number: float = - dt.datetime.now().timestamp()

    def __lt__(self, other: AlternativeInput) -> bool:
        return self.timestamp < other.timestamp

    def __eq__(self, other: AlternativeInput) -> bool:
        return self.timestamp == other.timestamp

    @property
    def timestamp(self) -> dt.datetime.timestamp :
        return cast(dt.datetime, self.current_time).timestamp()


def by_number(item: AlternativeInput) -> float :
    """Used for the Key Parameter in the sort method when looking to sort a list"""
    return cast(float, item.number)

@dataclass(order=True, frozen=True)
class AlternativeInputOrdered:
    staging_dir:str
    destination_dir: str
    current_time: dt.datetime = dt.datetime.now()

class RonaldContextManager():
    def __enter__(self):
        self.file = open(Path('./main.py'))
        return self.file
    def __exit__(self, one, two, three):
        self.file.close()
        print(f'file: {self.file} being closed')
    def __call__(self, *args: Any, **kwds: Any) -> Any:
        print(f'{self.__repr__()} being called being as an object')
        pass

class ObserverLogging(RonaldContextManager):
    def __enter__(self):
        mode = 'a' if Path(self._file_path).is_file() else 'w'
        self.file = open(Path(self._file_path), mode)
        self.file.write('\n') if mode == 'a' else None
        self.file.write(f'Serialization done with {self._sso._name} by {self._sso._user} at {dt.datetime.now()}')
        return self.file
    # TODO: see if this can be made a class method
    def __call__(self, observer: SheetSerializerObserver) -> None:
        self._sso = observer
        self._file_path = os.path.join(os.getcwd(), f'{observer._user}_{observer._name}_{observer._created_at.date()}')
if __name__ == "__main__":
    

    # argument should go python main.py <directory of excel sheets> <desired directory for pdfs>
    try :
        if len(sys.argv) != 3 or (not sys.argv[1] and not sys.argv[2]):
            raise MissingInputError
        anas_input = UserInput(base_dir=sys.argv[1], dest_dir=sys.argv[2])
    except IndexError as ex:
        print(f'the issue is {ex!r}')
        # raise ex
    finally:
        print(f'fix the following and try again')
    directory = create_data_directory(anas_input.dest_dir)
    print(directory)

    directory_of_excel_sheets = sys.argv[1]
    desired_directory_for_pdfs = sys.argv[2]

    # directory = desired_directory_for_pdfs + directory
    the_workbooks = glob.glob(f"{directory_of_excel_sheets}/*.xlsx")
    # the_sheets = glob.glob(f"excel_sheets/*.xlsx")
    # files = itemgetter(*[0])(the_sheets) not needed

    # file should be searched by suffix
    for wkbk in the_workbooks:
        book = openpyxl.load_workbook(f"{wkbk}", data_only=True)
        workbook_parser = WorkbookParser(workbook=book)
        workbook_parser.parser()

    exit(0)
    # the sheet name has the client's names
    book.sheetnames
    try:
        print('hi')
        raise SystemExit
    except ValueError as e:
        print(e)
    finally:
        print('caching before exit')
    for a in book.worksheets:
        print(a.title)

    # data range
    client1 = book.worksheets[1]
    mx_rw, mx_col = client1.max_row, client1.max_column

    # reading data (1-based ix) 
    # go to row then change col instead
    prev_val = None
    for row in range(1, mx_rw + 1): # starting with the first row (top)
        for col in range(mx_col, 0, -1): # starting with the last column
            val = client1.cell(row=row, column=col).value #value of the coordinates
            # if val:
            #     print(val)
            #     print(col,row)
            lw_cs_val = val.lower() if (type.val) == str else val # trnasofmring to lower case if datatype is str
            if re.match(r'.*date.*', lw_cs_val) and type(prev_val) == dt.datetime :# if date is in the current value and the prev_val (column to right)
                # cell_contents = prev_val.split(' ')                               # is of type datetime      # re.match(r'.*\d{4}', prev_val):
                # month = cell_contents[1]
                prev_val_month :dt.datetime = prev_val.strftime("%b")
                # insert code for creating pdf
            prev_val = val
            print(val)

    # I want to take each sheet serialized separately
    # then convert this file to pdf
    # https://stackoverflow.com/questions/57724345/print-excel-to-pdf-with-xlwings
    # client1.cell(row=9, column=8).value.strftime("%b")
    # re.match(r'.*date.*', client1.cell(row=6, column=7).value.lower())
    # re.match(r'.*date.*', 'dates')

    # use this to change the format of the sheets (ultimately take allow the option of taking arguments)
    # create a function that takes a sheet number of the xlsx and then create the pdf of it
    book = xw.Book("excel_sheets/2022.02.28 Test Invoice to automate.xlsx")
    sheet = book.sheets[1]

    pdf_path = os.path.join(os.getcwd(), 'file.pdf') 

    book.to_pdf(path=pdf_path, include=2)
