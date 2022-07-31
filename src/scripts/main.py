from excel_analytics import utils
import sys
import glob
import openpyxl
def main():
    print("Just testing my Python packaging skills :)")
if __name__ == "__main__":
    # argument should go python main.py <directory of excel sheets> <desired directory for pdfs>
    try :
        if len(sys.argv) != 3 or (not sys.argv[1] and not sys.argv[2]):
            raise utils.MissingInputError
        anas_input = utils.UserInput(base_dir=sys.argv[1], dest_dir=sys.argv[2])
    except IndexError as ex:
        print(f'the issue is {ex!r}')
        # raise ex
    finally:
        print(f'fix the following and try again')
    directory = utils.create_data_directory(anas_input.dest_dir)
    print(directory)

    directory_of_excel_sheets = sys.argv[1]
    desired_directory_for_pdfs = sys.argv[2]

    # directory = desired_directory_for_pdfs + directory
    the_workbooks = glob.glob(f"{directory_of_excel_sheets}/*.xlsx")
    # the_sheets = glob.glob(f"excel_sheets/*.xlsx")
    # files = itemgetter(*[0])(the_sheets) not needed

    # file should be searched by suffix
    for wkbk in the_workbooks:
        book = openpyxl.load_workbook(f"{wkbk}", data_only=True)
        workbook_parser = utils.WorkbookParser(workbook=book)
        workbook_parser.parser()

    exit(0)
    # the sheet name has the client's names
    book.sheetnames
    try:
        print('hi')
        raise SystemExit
    except ValueError as e:
        print(e)
    finally:
        print('caching before exit')
    for a in book.worksheets:
        print(a.title)

    # data range
    client1 = book.worksheets[1]
    mx_rw, mx_col = client1.max_row, client1.max_column

    # reading data (1-based ix) 
    # go to row then change col instead
    prev_val = None
    for row in range(1, mx_rw + 1): # starting with the first row (top)
        for col in range(mx_col, 0, -1): # starting with the last column
            val = client1.cell(row=row, column=col).value #value of the coordinates
            # if val:
            #     print(val)
            #     print(col,row)
            lw_cs_val = val.lower() if (type.val) == str else val # trnasofmring to lower case if datatype is str
            if re.match(r'.*date.*', lw_cs_val) and type(prev_val) == dt.datetime :# if date is in the current value and the prev_val (column to right)
                # cell_contents = prev_val.split(' ')                               # is of type datetime      # re.match(r'.*\d{4}', prev_val):
                # month = cell_contents[1]
                prev_val_month :dt.datetime = prev_val.strftime("%b")
                # insert code for creating pdf
            prev_val = val
            print(val)

    # I want to take each sheet serialized separately
    # then convert this file to pdf
    # https://stackoverflow.com/questions/57724345/print-excel-to-pdf-with-xlwings
    # client1.cell(row=9, column=8).value.strftime("%b")
    # re.match(r'.*date.*', client1.cell(row=6, column=7).value.lower())
    # re.match(r'.*date.*', 'dates')

    # use this to change the format of the sheets (ultimately take allow the option of taking arguments)
    # create a function that takes a sheet number of the xlsx and then create the pdf of it
    book = xw.Book("excel_sheets/2022.02.28 Test Invoice to automate.xlsx")
    sheet = book.sheets[1]

    pdf_path = os.path.join(os.getcwd(), 'file.pdf') 

    book.to_pdf(path=pdf_path, include=2)
